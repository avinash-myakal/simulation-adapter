# =========================================
# file: mm_final_energy_sim.py
# (unchanged logic; only the alias added at bottom)
# =========================================
from __future__ import annotations

import csv
import io
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
from typing import Dict, Iterable, List, Optional, Tuple

import boto3
from botocore.exceptions import ClientError
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

S3_ENDPOINT = os.environ.get("S3_ENDPOINT", "http://minio:9000")
S3_ACCESS_KEY = os.environ.get("S3_ACCESS_KEY", "minioadmin")
S3_SECRET_KEY = os.environ.get("S3_SECRET_KEY", "minioadmin")
S3_BUCKET = os.environ.get("S3_BUCKET", "mmstore")

S3_IDF_PREFIX = os.environ.get("S3_IDF_PREFIX", "output_idf_files/")
S3_EPW_PREFIX = os.environ.get("S3_EPW_PREFIX", "weather/epw/")

TOTAL_XLSX_KEY = "results_total_site_energy.xlsx"
PERAREA_XLSX_KEY = "results_site_energy_per_area.xlsx"

ENERGYPLUS_EXE = os.environ.get("ENERGYPLUS_EXE", "energyplus")
TIMEOUT_SEC = int(os.environ.get("RUN_TIMEOUT_SEC", "3600"))
KEEP_OUTPUTS = os.environ.get("KEEP_OUTPUTS", "false").lower() in ("1", "true", "yes")

YEARS_START = 2025
YEARS_END = 2084

IDF_NAME_RE = re.compile(
    r"^(RCP(?:8\.5|4\.5|2\.6))_([EFT](?:[12]?\d|2[0-8]))_(20[2-7]\d|208[0-4])\.idf$",
    re.IGNORECASE,
)
EPW_NAME_RE = re.compile(
    r"^Rotterdam_(RCP(?:85|45|26))_(20[2-7]\d|208[0-4])\.epw$",
    re.IGNORECASE,
)

def map_rcp_dot_to_nodot(rcp: str) -> str:
    rcp = rcp.upper().replace(" ", "")
    return rcp.replace("8.5", "85").replace("4.5", "45").replace("2.6", "26")

def parse_idf_filename(name: str) -> Tuple[str, str, int]:
    m = IDF_NAME_RE.match(name)
    if not m:
        raise ValueError(f"Unrecognized IDF filename: {name}")
    rcp_dot, aging, year_str = m.groups()
    return rcp_dot.upper(), aging.upper(), int(year_str)

def s3_client():
    return boto3.client(
        "s3",
        endpoint_url=S3_ENDPOINT,
        aws_access_key_id=S3_ACCESS_KEY,
        aws_secret_access_key=S3_SECRET_KEY,
    )

def _list_objects_all(c, bucket: str, prefix: str) -> Iterable[dict]:
    token = None
    while True:
        kwargs = {"Bucket": bucket, "Prefix": prefix}
        if token:
            kwargs["ContinuationToken"] = token
        resp = c.list_objects_v2(**kwargs)
        for obj in resp.get("Contents", []) or []:
            yield obj
        if not resp.get("IsTruncated"):
            break
        token = resp.get("NextContinuationToken")

def object_exists(c, bucket: str, key: str) -> bool:
    try:
        c.head_object(Bucket=bucket, Key=key)
        return True
    except ClientError as e:
        code = e.response.get("ResponseMetadata", {}).get("HTTPStatusCode")
        if code in (403, 404):
            return False
        raise

def download_to_file(c, bucket: str, key: str, local_path: str) -> None:
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    c.download_file(bucket, key, local_path)

def upload_bytes(c, bucket: str, key: str, data: bytes) -> None:
    c.put_object(Bucket=bucket, Key=key, Body=data)

def index_epws_s3(c, bucket: str, epw_prefix: str) -> Dict[Tuple[str, int], str]:
    idx: Dict[Tuple[str, int], str] = {}
    for obj in _list_objects_all(c, bucket, epw_prefix):
        key = obj["Key"]
        name = key.rsplit("/", 1)[-1]
        m = EPW_NAME_RE.match(name)
        if not m:
            continue
        rcp_nodot, year_str = m.groups()
        idx[(rcp_nodot.upper(), int(year_str))] = key
    return idx

def run_energyplus(idf_path: str, epw_path: str, outdir: str) -> None:
    cmd = [
        ENERGYPLUS_EXE,
        "--weather", epw_path,
        "--output-directory", outdir,
        "--readvars", "--expandobjects",
        idf_path,
    ]
    proc = subprocess.run(
        cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=TIMEOUT_SEC
    )
    if proc.returncode != 0:
        raise RuntimeError(
            f"EnergyPlus failed ({proc.returncode}) for:\n  {os.path.basename(idf_path)}\nSTDOUT:\n{proc.stdout}\nSTDERR:\n{proc.stderr}"
        )

def _safe_float(s: str) -> Optional[float]:
    try:
        return float(str(s).replace(",", "").strip())
    except Exception:
        return None

def fetch_from_sql(sql_path: str) -> Tuple[Optional[float], Optional[float]]:
    if not os.path.exists(sql_path):
        return None, None
    con = sqlite3.connect(sql_path)
    cur = con.cursor()
    try:
        cur.execute(
            """
            SELECT Value
            FROM TabularDataWithStrings
            WHERE lower(TableName) LIKE 'site and source energy%%'
              AND lower(RowName) = 'total site energy'
              AND lower(ColumnName) LIKE 'total energy%%'
            """
        )
        row = cur.fetchone()
        total_site = _safe_float(row[0]) if row else None

        cur.execute(
            """
            SELECT Value
            FROM TabularDataWithStrings
            WHERE lower(TableName) LIKE 'site and source energy%%'
              AND lower(RowName) = 'total site energy'
              AND lower(ColumnName) LIKE 'energy per total building area%%'
            """
        )
        row = cur.fetchone()
        per_area = _safe_float(row[0]) if row else None
        return total_site, per_area
    finally:
        con.close()

def _fetch_from_tabular_text(path: str, delimiter: str) -> Tuple[Optional[float], Optional[float]]:
    if not os.path.exists(path):
        return None, None
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        rows: List[List[str]] = [[c.strip() for c in r] for r in csv.reader(f, delimiter=delimiter)]
    for i, r in enumerate(rows):
        if any("annual building utility performance summary" in c.lower() for c in r):
            for j in range(i + 1, min(i + 40, len(rows))):
                if any("site and source energy" in c.lower() for c in rows[j]):
                    header_idx = None
                    for k in range(j + 1, min(j + 15, len(rows))):
                        hdr_lower = [c.lower() for c in rows[k]]
                        if any("total energy" in c for c in hdr_lower) and any(
                            "energy per total building area" in c for c in hdr_lower
                        ):
                            header_idx = k
                            break
                    if header_idx is None:
                        continue
                    header = [h.lower() for h in rows[header_idx]]
                    try:
                        col_total = next(i2 for i2, c in enumerate(header) if "total energy" in c)
                        col_per_area = next(i2 for i2, c in enumerate(header) if "energy per total building area" in c)
                    except StopIteration:
                        continue
                    for r2 in rows[header_idx + 1 :]:
                        if not r2 or all(c == "" for c in r2):
                            break
                        if r2[0].strip().lower() == "total site energy":
                            total_site = _safe_float(r2[col_total]) if col_total < len(r2) else None
                            per_area = _safe_float(r2[col_per_area]) if col_per_area < len(r2) else None
                            return total_site, per_area
    return None, None

def fetch_metrics_any(outdir: str) -> Tuple[Optional[float], Optional[float], str]:
    sql_path = os.path.join(outdir, "eplusout.sql")
    if os.path.exists(sql_path):
        t, p = fetch_from_sql(sql_path)
        if t is not None or p is not None:
            return t, p, "sql"
    tab_path = os.path.join(outdir, "eplustbl.tab")
    t, p = _fetch_from_tabular_text(tab_path, "\t")
    if t is not None or p is not None:
        return t, p, "tab"
    csv_path = os.path.join(outdir, "eplustbl.csv")
    t, p = _fetch_from_tabular_text(csv_path, ",")
    if t is not None or p is not None:
        return t, p, "csv"
    return None, None, "none"

def _init_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws["A1"] = "Year"; ws["A2"] = ""
    for i, year in enumerate(range(YEARS_START, YEARS_END + 1), start=3):
        ws.cell(row=i, column=1, value=year)
    return wb

def _download_workbook_or_create(c, bucket: str, key: str) -> Workbook:
    if object_exists(c, bucket, key):
        obj = c.get_object(Bucket=bucket, Key=key)
        data = obj["Body"].read()
        return load_workbook(io.BytesIO(data))
    return _init_workbook()

def _upload_workbook(c, bucket: str, key: str, wb: Workbook) -> None:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    upload_bytes(c, bucket, key, buf.read())

def _col_for_pair(ws, climate: str, aging: str) -> int:
    max_col = ws.max_column or 1
    for col in range(2, max_col + 1):
        c1 = (ws.cell(row=1, column=col).value or "").strip()
        c2 = (ws.cell(row=2, column=col).value or "").strip()
        if c1 == climate and c2 == aging:
            return col
    col = max_col + 1 if max_col >= 2 else 2
    ws.cell(row=1, column=col, value=climate)
    ws.cell(row=2, column=col, value=aging)
    ws.column_dimensions[get_column_letter(col)].width = 16
    return col

def _row_for_year(year: int) -> int:
    if year < YEARS_START or year > YEARS_END:
        raise ValueError(f"Year {year} outside [{YEARS_START}, {YEARS_END}]")
    return 3 + (year - YEARS_START)

def write_result_to_s3(c, bucket: str, key: str, climate: str, aging: str, year: int, value: Optional[float]) -> None:
    if value is None:
        return  # why: match original behavior; don't write zeros
    wb = _download_workbook_or_create(c, bucket, key)
    ws = wb.active
    col = _col_for_pair(ws, climate, aging)
    row = _row_for_year(year)
    ws.cell(row=row, column=col, value=value)
    _upload_workbook(c, bucket, key, wb)
    wb.close()

def run_energy_simulation_from_env() -> Dict[str, object]:
    c = s3_client()
    epw_index = index_epws_s3(c, S3_BUCKET, S3_EPW_PREFIX)
    if not epw_index:
        raise SystemExit(f"No EPWs under s3://{S3_BUCKET}/{S3_EPW_PREFIX}")

    idf_objs = [o for o in _list_objects_all(c, S3_BUCKET, S3_IDF_PREFIX) if o["Key"].lower().endswith(".idf")]
    if not idf_objs:
        raise SystemExit(f"No IDFs under s3://{S3_BUCKET}/{S3_IDF_PREFIX}")

    successes = failures = 0
    processed: List[str] = []

    for i, obj in enumerate(sorted(idf_objs, key=lambda x: x["Key"]), start=1):
        idf_key = obj["Key"]
        name = idf_key.rsplit("/", 1)[-1]
        run_dir = None
        try:
            rcp_dot, aging, year = parse_idf_filename(name)
            rcp_nodot = map_rcp_dot_to_nodot(rcp_dot)
            epw_key = epw_index.get((rcp_nodot, year))
            if not epw_key:
                print(f"[WARN] No EPW for {name} → ({rcp_nodot}, {year}); skipping.")
                failures += 1
                continue

            run_dir = tempfile.mkdtemp(prefix="eplus_run_")
            idf_path = os.path.join(run_dir, name)
            epw_name = epw_key.rsplit("/", 1)[-1]
            epw_path = os.path.join(run_dir, epw_name)

            download_to_file(c, S3_BUCKET, idf_key, idf_path)
            download_to_file(c, S3_BUCKET, epw_key, epw_path)

            print(f"[{i}/{len(idf_objs)}] {name} × {epw_name}")
            run_energyplus(idf_path, epw_path, run_dir)

            total_site, per_area, src = fetch_metrics_any(run_dir)
            if total_site is None and per_area is None:
                raise RuntimeError("Failed to extract metrics; ensure ABUPS tables are produced.")

            write_result_to_s3(c, S3_BUCKET, TOTAL_XLSX_KEY, rcp_nodot, aging, year, total_site)
            write_result_to_s3(c, S3_BUCKET, PERAREA_XLSX_KEY, rcp_nodot, aging, year, per_area)

            successes += 1
            processed.append(idf_key)
            print(f"    -> OK (src={src}, total_site={total_site}, per_area={per_area})")
        except Exception as e:
            failures += 1
            print(f"[ERROR] {name}: {e}")
        finally:
            if run_dir and os.path.isdir(run_dir) and not KEEP_OUTPUTS:
                shutil.rmtree(run_dir, ignore_errors=True)

    return {
        "bucket": S3_BUCKET,
        "idf_prefix": S3_IDF_PREFIX.rstrip("/"),
        "epw_prefix": S3_EPW_PREFIX.rstrip("/"),
        "results_total_key": TOTAL_XLSX_KEY,
        "results_per_area_key": PERAREA_XLSX_KEY,
        "processed": processed[:10],
        "counts": {"success": successes, "failed": failures, "total": len(idf_objs)},
    }

def main() -> None:
    print(run_energy_simulation_from_env())

if __name__ == "__main__":
    main()

# ---------- compatibility alias (fix for adapter name mismatch) ----------
# Why: some adapters call this legacy name; keep both without duplicating logic.
run_energy_model_from_env = run_energy_simulation_from_env
